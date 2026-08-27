# """
# server.py — Unified AM Data Pipeline Server
# Runs BOTH the versioned REST API (/api/v1/) AND serves all dashboard HTML files.
# Single command: python server.py
# """

# import os
# import uvicorn
# from fastapi import Depends, HTTPException
# from fastapi.responses import FileResponse, HTMLResponse
# from fastapi.staticfiles import StaticFiles
# from fastapi.middleware.cors import CORSMiddleware
# from sqlalchemy import text
# from sqlalchemy.orm import Session, configure_mappers
# from datetime import datetime
# from fastapi import UploadFile, File, Form
# import io

# # ── Step 1: import ALL database models first ──────────────────────────────
# from am_data_pipeline_postgres import (
#     engine, get_db,
#     Experiment, ProcessParameter, GeometryData,
#     QualityMetric, SensorData, MLFeature
# )

# # ── Step 2: resolve all ORM relationships immediately ─────────────────────
# try:
#     configure_mappers()
#     print("[OK]   SQLAlchemy mappers configured")
# except Exception as e:
#     print(f"[WARN] Mapper warning (non-fatal): {e}")

# # ── Step 3: load the core versioned API app ───────────────────────────────
# from restful_api import app

# # ── Step 4: add CORS so HTML files opened from file:// also work ──────────
# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

# # ── Step 5: load extension modules safely ────────────────────────────────
# modules_loaded = []
# modules_failed = []

# for module_name in [
#     "version_control_api",
#     "data_quality_api",
#     "data_analysis_api",
#     "ml_api",
#     "security_privacy",
#     "backup_recovery",
#     "collaboration_sharing",
# ]:
#     try:
#         __import__(module_name)
#         modules_loaded.append(module_name)
#     except Exception as e:
#         modules_failed.append((module_name, str(e)[:120]))

# print("\n=== Module load report ===")
# for m in modules_loaded:
#     print(f"  [OK]   {m}")
# for m, err in modules_failed:
#     print(f"  [SKIP] {m}")
#     print(f"         {err}")
# print("==========================\n")

# # ── Step 6: register ALL routes ──────────────────────────────────────────

# # Health check (versioned)
# @app.get("/api/v1/health", tags=["Health"])
# def health_check(db: Session = Depends(get_db)):
#     try:
#         db.execute(text("SELECT 1"))
#         return {
#             "status": "healthy",
#             "database": "connected",
#             "modules_loaded": modules_loaded,
#             "modules_skipped": [m for m, _ in modules_failed],
#             "timestamp": datetime.now().isoformat()
#         }
#     except Exception as e:
#         raise HTTPException(status_code=503, detail=f"Database error: {str(e)}")

# # Legacy unversioned routes — these make am_dashboard.html work
# # because that file calls /api/experiments (no /v1/)
# @app.get("/api/experiments")
# def legacy_list_experiments(
#     material_type: str = None, # type: ignore
#     limit: int = 100,
#     offset: int = 0,
#     db: Session = Depends(get_db)
# ):
#     """Legacy unversioned route — forwards to v1 logic"""
#     from sqlalchemy import func
#     query = db.query(Experiment)
#     if material_type:
#         query = query.filter(Experiment.material_type == material_type)
#     experiments = query.order_by(
#         Experiment.build_date.desc()
#     ).offset(offset).limit(limit).all()

#     result = []
#     for exp in experiments:
#         result.append({
#             "experiment_id": exp.experiment_id,
#             "experiment_name": exp.experiment_name,
#             "material_type": exp.material_type,
#             "status": exp.status,
#             "build_date": exp.build_date.isoformat() if exp.build_date else None, # type: ignore
#             "process_parameters": {
#                 "layer_height": exp.process_parameters.layer_height,
#                 "print_speed": exp.process_parameters.print_speed,
#                 "nozzle_temperature": exp.process_parameters.nozzle_temperature,
#                 "infill_percentage": exp.process_parameters.infill_percentage,
#             } if exp.process_parameters else None,
#             "quality_metrics": {
#                 "tensile_strength_mpa": exp.quality_metrics.tensile_strength_mpa,
#                 "surface_roughness_um": exp.quality_metrics.surface_roughness_um,
#                 "porosity_percent": exp.quality_metrics.porosity_percent,
#             } if exp.quality_metrics else None,
#         })
#     return result

# @app.get("/api/experiments/{experiment_id}")
# def legacy_get_experiment(experiment_id: str, db: Session = Depends(get_db)):
#     """Legacy unversioned route"""
#     exp = db.query(Experiment).filter(
#         Experiment.experiment_id == experiment_id
#     ).first()
#     if not exp:
#         raise HTTPException(status_code=404, detail="Experiment not found")
#     return {
#         "experiment_id": exp.experiment_id,
#         "experiment_name": exp.experiment_name,
#         "material_type": exp.material_type,
#         "status": exp.status,
#         "build_date": exp.build_date.isoformat() if exp.build_date else None, # type: ignore
#         "process_parameters": exp.process_parameters.__dict__ if exp.process_parameters else None,
#         "quality_metrics": exp.quality_metrics.__dict__ if exp.quality_metrics else None,
#         "geometry_data": exp.geometry_data.__dict__ if exp.geometry_data else None,
#     }

# @app.post("/api/experiments")
# def legacy_create_experiment(experiment: dict, db: Session = Depends(get_db)):
#     """Legacy unversioned route — redirects to v1 create"""
#     from restful_api import create_experiment, ExperimentCreateModel
#     return create_experiment(ExperimentCreateModel(**experiment), db)

# @app.get("/api/analytics/summary")
# def legacy_analytics_summary(db: Session = Depends(get_db)):
#     """Legacy unversioned route — makes am_dashboard.html analytics work"""
#     from sqlalchemy import func
#     total = db.query(Experiment).count()
#     material_dist = dict(
#         db.query(Experiment.material_type, func.count(Experiment.experiment_id))
#         .group_by(Experiment.material_type).all() # type: ignore
#     ) # type: ignore
#     avg_quality = db.query(
#         func.avg(QualityMetric.tensile_strength_mpa),
#         func.avg(QualityMetric.surface_roughness_um),
#         func.avg(QualityMetric.porosity_percent)
#     ).first()
#     param_ranges = db.query(
#         func.min(ProcessParameter.nozzle_temperature),
#         func.max(ProcessParameter.nozzle_temperature),
#         func.avg(ProcessParameter.nozzle_temperature),
#         func.min(ProcessParameter.print_speed),
#         func.max(ProcessParameter.print_speed),
#         func.avg(ProcessParameter.print_speed)
#     ).first()
#     return {
#         "total_experiments": total,
#         "material_distribution": {k: v for k, v in material_dist.items() if k},
#         "average_quality_metrics": {
#             "tensile_strength_mpa": float(avg_quality[0]) if avg_quality[0] else None, # type: ignore
#             "surface_roughness_um": float(avg_quality[1]) if avg_quality[1] else None, # type: ignore
#             "porosity_percent": float(avg_quality[2]) if avg_quality[2] else None, # type: ignore
#         },
#         "process_parameter_ranges": {
#             "nozzle_temperature": {
#                 "min": float(param_ranges[0]) if param_ranges[0] else None, # type: ignore
#                 "max": float(param_ranges[1]) if param_ranges[1] else None, # type: ignore
#                 "avg": float(param_ranges[2]) if param_ranges[2] else None, # type: ignore
#             },
#             "print_speed": {
#                 "min": float(param_ranges[3]) if param_ranges[3] else None, # type: ignore
#                 "max": float(param_ranges[4]) if param_ranges[4] else None, # type: ignore
#                 "avg": float(param_ranges[5]) if param_ranges[5] else None, # type: ignore
#             }
#         }
#     }

# @app.get("/api/export/ml-dataset")
# def legacy_export(
#     format: str = "csv",
#     material_type: str = None, # type: ignore
#     db: Session = Depends(get_db)
# ):
#     """Legacy unversioned export route"""
#     import pandas as pd
#     from io import BytesIO
#     from fastapi.responses import StreamingResponse
#     query = db.query(
#         Experiment.experiment_id, Experiment.material_type,
#         ProcessParameter.layer_height, ProcessParameter.print_speed,
#         ProcessParameter.nozzle_temperature, ProcessParameter.infill_percentage,
#         QualityMetric.tensile_strength_mpa, QualityMetric.surface_roughness_um,
#         QualityMetric.porosity_percent
#     ).outerjoin(ProcessParameter).outerjoin(QualityMetric)
#     if material_type:
#         query = query.filter(Experiment.material_type == material_type)
#     df = pd.read_sql(query.statement, db.bind) # type: ignore
#     if format == "csv":
#         output = BytesIO()
#         df.to_csv(output, index=False)
#         output.seek(0)
#         return StreamingResponse(output, media_type="text/csv",
#             headers={"Content-Disposition": "attachment; filename=am_ml_dataset.csv"})
#     elif format == "parquet":
#         output = BytesIO()
#         df.to_parquet(output, index=False)
#         output.seek(0)
#         return StreamingResponse(output, media_type="application/octet-stream",
#             headers={"Content-Disposition": "attachment; filename=am_ml_dataset.parquet"})
#     return df.to_dict(orient="records")

# # Dashboard HTML routes
# @app.get("/")
# def root_dashboard():
#     return FileResponse("am_dashboard.html")

# @app.get("/dashboard/basic")
# def basic_dashboard():
#     return FileResponse("am_dashboard.html")

# @app.get("/dashboard/advanced")
# def advanced_dashboard():
#     return FileResponse("am_dashboard_advanced.html")

# @app.get("/upload")
# def upload_page():
#     return FileResponse("upload_dashboard.html")

# @app.post("/api/upload/csv")
# async def upload_file(
#     file: UploadFile = File(...),
#     db: Session = Depends(get_db)
# ):
#     """
#     Accept CSV, TSV, TXT, XLSX, XLS, JSON or Parquet uploads
#     and ingest all rows into the database.
#     """
#     from am_data_pipeline_postgres import (
#         ProcessParameter, GeometryData, QualityMetric, compute_ml_features
#     )
#     import tempfile, os

#     content  = await file.read()
#     filename = file.filename.lower()

#     # ── parse by extension ─────────────────────────────────────────────────
#     try:
#         if filename.endswith(".xlsx") or filename.endswith(".xls"):
#             import openpyxl, io
#             wb   = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
#             ws   = wb.active
#             rows_raw = list(ws.iter_rows(values_only=True))
#             if not rows_raw:
#                 return {"error": "Empty Excel file"}
#             header = [
#                 str(c) if c is not None else f"col_{i}"
#                 for i, c in enumerate(rows_raw[0])
#             ]
#             raw_rows = [
#                 {header[i]: (str(v) if v is not None else "")
#                  for i, v in enumerate(row)}
#                 for row in rows_raw[1:]
#                 if not all(v is None for v in row)
#             ]

#         elif filename.endswith(".json"):
#             import json
#             data = json.loads(content.decode("utf-8"))
#             if isinstance(data, list):
#                 raw_rows = [
#                     {k: (str(v) if v is not None else "")
#                      for k, v in item.items()}
#                     for item in data if isinstance(item, dict)
#                 ]
#             elif isinstance(data, dict):
#                 raw_rows = []
#                 for val in data.values():
#                     if isinstance(val, list) and val:
#                         raw_rows = [
#                             {k: (str(v) if v is not None else "")
#                              for k, v in item.items()}
#                             for item in val if isinstance(item, dict)
#                         ]
#                         break
#                 if not raw_rows:
#                     raw_rows = [{k: str(v) for k, v in data.items()}]
#             else:
#                 return {"error": "JSON must be an array or object"}

#         elif filename.endswith(".parquet"):
#             import pandas as pd, io
#             df       = pd.read_parquet(io.BytesIO(content))
#             df       = df.where(df.notna(), other="")
#             raw_rows = df.astype(str).to_dict(orient="records")

#         elif filename.endswith(".tsv") or filename.endswith(".txt"):
#             import csv, io as _io
#             text     = content.decode("utf-8-sig")
#             sample   = text[:4096]
#             delim    = "\t" if sample.count("\t") > sample.count(",") else ","
#             reader   = csv.DictReader(_io.StringIO(text), delimiter=delim)
#             raw_rows = [dict(row) for row in reader]

#         else:
#             # Default: CSV
#             import csv, io as _io
#             text     = content.decode("utf-8-sig")
#             reader   = csv.DictReader(_io.StringIO(text))
#             raw_rows = [dict(row) for row in reader]

#     except Exception as e:
#         return {"error": f"Could not parse file: {str(e)}"}

#     # ── column normalisation ───────────────────────────────────────────────
#     CMAP = {
#         "id":"experiment_id","exp_id":"experiment_id",
#         "sample_id":"experiment_id","specimen_id":"experiment_id",
#         "name":"experiment_name","material":"material_type",
#         "mat":"material_type","alloy":"material_type",
#         "batch":"material_batch","layer_thickness":"layer_height",
#         "scan_speed":"print_speed","laser_speed":"print_speed",
#         "nozzle_temp":"nozzle_temperature",
#         "extrusion_temp":"nozzle_temperature",
#         "platform_temp":"bed_temperature",
#         "infill":"infill_percentage","tensile":"tensile_strength_mpa",
#         "uts":"tensile_strength_mpa","roughness":"surface_roughness_um",
#         "ra":"surface_roughness_um","porosity":"porosity_percent",
#         "density":"density_g_per_cm3","elongation":"elongation_percent",
#         "hardness":"hardness_hb","defects":"defect_count",
#     }

#     def norm(row):
#         out = {}
#         for k, v in row.items():
#             clean = str(k).strip().lower().replace(" ","_").replace("-","_")
#             out[CMAP.get(clean, clean)] = v
#         return out

#     def pf(v):
#         try:
#             return float(v) if str(v).strip() not in (
#                 "","None","N/A","nan","NaN","-","null") else None
#         except: return None

#     def pi(v):
#         try:
#             return int(float(v)) if str(v).strip() not in (
#                 "","None","N/A") else None
#         except: return None

#     # ── insert rows ────────────────────────────────────────────────────────
#     created = skipped = failed = 0
#     errors  = []

#     for idx, raw in enumerate(raw_rows, 1):
#         row    = norm(raw)
#         exp_id = str(row.get("experiment_id") or "").strip() or f"UP-{idx:05d}"

#         existing = db.query(Experiment).filter(
#             Experiment.experiment_id == exp_id
#         ).first()
#         if existing:
#             skipped += 1
#             continue

#         try:
#             db.add(Experiment(
#                 experiment_id   = exp_id,
#                 experiment_name = str(row.get("experiment_name") or f"Upload {idx}"),
#                 material_type   = row.get("material_type")  or None,
#                 material_batch  = row.get("material_batch") or None,
#                 build_platform  = row.get("build_platform") or None,
#                 build_date      = row.get("build_date")      or None,
#                 operator        = row.get("operator")        or None,
#                 status          = row.get("status")          or "completed",
#                 notes           = row.get("notes")           or None,
#             ))
#             db.add(ProcessParameter(
#                 experiment_id     = exp_id,
#                 layer_height      = pf(row.get("layer_height")),
#                 print_speed       = pf(row.get("print_speed")),
#                 nozzle_temperature= pf(row.get("nozzle_temperature")),
#                 bed_temperature   = pf(row.get("bed_temperature")),
#                 infill_percentage = pf(row.get("infill_percentage")),
#                 infill_pattern    = row.get("infill_pattern") or None,
#                 shell_count       = pi(row.get("shell_count")),
#                 print_time_hours  = pf(row.get("print_time_hours")),
#             ))
#             db.add(GeometryData(
#                 experiment_id    = exp_id,
#                 volume_mm3       = pf(row.get("volume_mm3")),
#                 surface_area_mm2 = pf(row.get("surface_area_mm2")),
#                 bounding_box_x   = pf(row.get("bounding_box_x")),
#                 bounding_box_y   = pf(row.get("bounding_box_y")),
#                 bounding_box_z   = pf(row.get("bounding_box_z")),
#             ))
#             db.add(QualityMetric(
#                 experiment_id       = exp_id,
#                 tensile_strength_mpa= pf(row.get("tensile_strength_mpa")),
#                 yield_strength_mpa  = pf(row.get("yield_strength_mpa")),
#                 elongation_percent  = pf(row.get("elongation_percent")),
#                 surface_roughness_um= pf(row.get("surface_roughness_um")),
#                 porosity_percent    = pf(row.get("porosity_percent")),
#                 density_g_per_cm3   = pf(row.get("density_g_per_cm3")),
#                 hardness_hb         = pf(row.get("hardness_hb")),
#                 defect_count        = pi(row.get("defect_count")),
#             ))
#             db.commit()
#             compute_ml_features(db, exp_id)
#             created += 1
#         except Exception as e:
#             db.rollback()
#             failed += 1
#             errors.append({"row": idx, "id": exp_id, "error": str(e)[:120]})

#     return {
#         "filename":   file.filename,
#         "format":     os.path.splitext(file.filename)[1].upper().lstrip("."), # type: ignore
#         "total_rows": len(raw_rows),
#         "created":    created,
#         "skipped":    skipped,
#         "failed":     failed,
#         "errors":     errors[:10]
#     }




# # ── Step 7: startup banner ────────────────────────────────────────────────
# if __name__ == "__main__":
#     print("=" * 60)
#     print("  AM Data Pipeline — Unified Server")
#     print("=" * 60)
#     print("  Basic dashboard   : http://localhost:8000")
#     print("  Advanced dashboard: http://localhost:8000/dashboard/advanced")
#     print("  API docs (Swagger): http://localhost:8000/api/v1/docs")
#     print("  Health check      : http://localhost:8000/api/v1/health")
#     print("  Analytics JSON    : http://localhost:8000/api/analytics/summary")
#     print("=" * 60 + "\n")
#     uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)




"""
server.py — Unified AM Data Pipeline Server
Runs BOTH the versioned REST API (/api/v1/) AND serves all dashboard HTML files.
Single command: python server.py
"""

import os
import uvicorn
from fastapi import Depends, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import text
from sqlalchemy.orm import Session, configure_mappers
from datetime import datetime
from fastapi import UploadFile, File, Form
import io

# ── Step 1: import ALL database models first ──────────────────────────────
from am_data_pipeline_postgres import (
    engine, get_db,
    Experiment, ProcessParameter, GeometryData,
    QualityMetric, SensorData, MLFeature
)

# Directory this server.py lives in — so HTML files are found no matter
# what working directory the cloud host launches the process from.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _html(filename: str) -> str:
    """Return an absolute path to an HTML file next to this script."""
    return os.path.join(BASE_DIR, filename)

# ── Step 2: resolve all ORM relationships immediately ─────────────────────
try:
    configure_mappers()
    print("[OK]   SQLAlchemy mappers configured")
except Exception as e:
    print(f"[WARN] Mapper warning (non-fatal): {e}")

# ── Step 3: load the core versioned API app ───────────────────────────────
from restful_api import app

# ── Step 4: add CORS so HTML files opened from file:// also work ──────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Step 5: load extension modules safely ────────────────────────────────
modules_loaded = []
modules_failed = []

for module_name in [
    "version_control_api",
    "data_quality_api",
    "data_analysis_api",
    "ml_api",
    "security_privacy",
    "backup_recovery",
    "collaboration_sharing",
]:
    try:
        __import__(module_name)
        modules_loaded.append(module_name)
    except Exception as e:
        modules_failed.append((module_name, str(e)[:120]))

print("\n=== Module load report ===")
for m in modules_loaded:
    print(f"  [OK]   {m}")
for m, err in modules_failed:
    print(f"  [SKIP] {m}")
    print(f"         {err}")
print("==========================\n")

# ── Step 6: register ALL routes ──────────────────────────────────────────

# Health check (versioned)
@app.get("/api/v1/health", tags=["Health"])
def health_check(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {
            "status": "healthy",
            "database": "connected",
            "modules_loaded": modules_loaded,
            "modules_skipped": [m for m, _ in modules_failed],
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database error: {str(e)}")

# Legacy unversioned routes — these make am_dashboard.html work
# because that file calls /api/experiments (no /v1/)
@app.get("/api/experiments")
def legacy_list_experiments(
    material_type: str = None, # type: ignore
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Legacy unversioned route — forwards to v1 logic"""
    from sqlalchemy import func
    query = db.query(Experiment)
    if material_type:
        query = query.filter(Experiment.material_type == material_type)
    experiments = query.order_by(
        Experiment.build_date.desc()
    ).offset(offset).limit(limit).all()

    result = []
    for exp in experiments:
        result.append({
            "experiment_id": exp.experiment_id,
            "experiment_name": exp.experiment_name,
            "material_type": exp.material_type,
            "status": exp.status,
            "build_date": exp.build_date.isoformat() if exp.build_date else None, # type: ignore
            "process_parameters": {
                "layer_height": exp.process_parameters.layer_height,
                "print_speed": exp.process_parameters.print_speed,
                "nozzle_temperature": exp.process_parameters.nozzle_temperature,
                "infill_percentage": exp.process_parameters.infill_percentage,
            } if exp.process_parameters else None,
            "quality_metrics": {
                "tensile_strength_mpa": exp.quality_metrics.tensile_strength_mpa,
                "surface_roughness_um": exp.quality_metrics.surface_roughness_um,
                "porosity_percent": exp.quality_metrics.porosity_percent,
            } if exp.quality_metrics else None,
        })
    return result

@app.get("/api/experiments/{experiment_id}")
def legacy_get_experiment(experiment_id: str, db: Session = Depends(get_db)):
    """Legacy unversioned route"""
    exp = db.query(Experiment).filter(
        Experiment.experiment_id == experiment_id
    ).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Experiment not found")
    return {
        "experiment_id": exp.experiment_id,
        "experiment_name": exp.experiment_name,
        "material_type": exp.material_type,
        "status": exp.status,
        "build_date": exp.build_date.isoformat() if exp.build_date else None, # type: ignore
        "process_parameters": exp.process_parameters.__dict__ if exp.process_parameters else None,
        "quality_metrics": exp.quality_metrics.__dict__ if exp.quality_metrics else None,
        "geometry_data": exp.geometry_data.__dict__ if exp.geometry_data else None,
    }

@app.post("/api/experiments")
def legacy_create_experiment(experiment: dict, db: Session = Depends(get_db)):
    """Legacy unversioned route — redirects to v1 create"""
    from restful_api import create_experiment, ExperimentCreateModel
    return create_experiment(ExperimentCreateModel(**experiment), db)

@app.get("/api/analytics/summary")
def legacy_analytics_summary(db: Session = Depends(get_db)):
    """Legacy unversioned route — makes am_dashboard.html analytics work"""
    from sqlalchemy import func
    total = db.query(Experiment).count()
    material_dist = dict(
        db.query(Experiment.material_type, func.count(Experiment.experiment_id))
        .group_by(Experiment.material_type).all() # type: ignore
    ) # type: ignore
    avg_quality = db.query(
        func.avg(QualityMetric.tensile_strength_mpa),
        func.avg(QualityMetric.surface_roughness_um),
        func.avg(QualityMetric.porosity_percent)
    ).first()
    param_ranges = db.query(
        func.min(ProcessParameter.nozzle_temperature),
        func.max(ProcessParameter.nozzle_temperature),
        func.avg(ProcessParameter.nozzle_temperature),
        func.min(ProcessParameter.print_speed),
        func.max(ProcessParameter.print_speed),
        func.avg(ProcessParameter.print_speed)
    ).first()
    return {
        "total_experiments": total,
        "material_distribution": {k: v for k, v in material_dist.items() if k},
        "average_quality_metrics": {
            "tensile_strength_mpa": float(avg_quality[0]) if avg_quality[0] else None, # type: ignore
            "surface_roughness_um": float(avg_quality[1]) if avg_quality[1] else None, # type: ignore
            "porosity_percent": float(avg_quality[2]) if avg_quality[2] else None, # type: ignore
        },
        "process_parameter_ranges": {
            "nozzle_temperature": {
                "min": float(param_ranges[0]) if param_ranges[0] else None, # type: ignore
                "max": float(param_ranges[1]) if param_ranges[1] else None, # type: ignore
                "avg": float(param_ranges[2]) if param_ranges[2] else None, # type: ignore
            },
            "print_speed": {
                "min": float(param_ranges[3]) if param_ranges[3] else None, # type: ignore
                "max": float(param_ranges[4]) if param_ranges[4] else None, # type: ignore
                "avg": float(param_ranges[5]) if param_ranges[5] else None, # type: ignore
            }
        }
    }

@app.get("/api/export/ml-dataset")
def legacy_export(
    format: str = "csv",
    material_type: str = None, # type: ignore
    db: Session = Depends(get_db)
):
    """Legacy unversioned export route"""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    query = db.query(
        Experiment.experiment_id, Experiment.material_type,
        ProcessParameter.layer_height, ProcessParameter.print_speed,
        ProcessParameter.nozzle_temperature, ProcessParameter.infill_percentage,
        QualityMetric.tensile_strength_mpa, QualityMetric.surface_roughness_um,
        QualityMetric.porosity_percent
    ).outerjoin(ProcessParameter).outerjoin(QualityMetric)
    if material_type:
        query = query.filter(Experiment.material_type == material_type)
    df = pd.read_sql(query.statement, db.bind) # type: ignore
    if format == "csv":
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return StreamingResponse(output, media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=am_ml_dataset.csv"})
    elif format == "parquet":
        output = BytesIO()
        df.to_parquet(output, index=False)
        output.seek(0)
        return StreamingResponse(output, media_type="application/octet-stream",
            headers={"Content-Disposition": "attachment; filename=am_ml_dataset.parquet"})
    return df.to_dict(orient="records")

# Dashboard HTML routes
@app.get("/")
def root_dashboard():
    return FileResponse(_html("am_dashboard.html"))

@app.get("/dashboard/basic")
def basic_dashboard():
    return FileResponse(_html("am_dashboard.html"))

@app.get("/dashboard/advanced")
def advanced_dashboard():
    return FileResponse(_html("am_dashboard_advanced.html"))

@app.get("/upload")
def upload_page():
    return FileResponse(_html("upload_dashboard.html"))

@app.post("/api/upload/csv")
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Accept CSV, TSV, TXT, XLSX, XLS, JSON or Parquet uploads
    and ingest all rows into the database.
    """
    from am_data_pipeline_postgres import (
        ProcessParameter, GeometryData, QualityMetric, compute_ml_features
    )
    import tempfile, os

    content  = await file.read()
    filename = file.filename.lower()

    # ── parse by extension ─────────────────────────────────────────────────
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl, io
            wb   = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws   = wb.active
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw:
                return {"error": "Empty Excel file"}
            header = [
                str(c) if c is not None else f"col_{i}"
                for i, c in enumerate(rows_raw[0])
            ]
            raw_rows = [
                {header[i]: (str(v) if v is not None else "")
                 for i, v in enumerate(row)}
                for row in rows_raw[1:]
                if not all(v is None for v in row)
            ]

        elif filename.endswith(".json"):
            import json
            data = json.loads(content.decode("utf-8"))
            if isinstance(data, list):
                raw_rows = [
                    {k: (str(v) if v is not None else "")
                     for k, v in item.items()}
                    for item in data if isinstance(item, dict)
                ]
            elif isinstance(data, dict):
                raw_rows = []
                for val in data.values():
                    if isinstance(val, list) and val:
                        raw_rows = [
                            {k: (str(v) if v is not None else "")
                             for k, v in item.items()}
                            for item in val if isinstance(item, dict)
                        ]
                        break
                if not raw_rows:
                    raw_rows = [{k: str(v) for k, v in data.items()}]
            else:
                return {"error": "JSON must be an array or object"}

        elif filename.endswith(".parquet"):
            import pandas as pd, io
            df       = pd.read_parquet(io.BytesIO(content))
            df       = df.where(df.notna(), other="")
            raw_rows = df.astype(str).to_dict(orient="records")

        elif filename.endswith(".tsv") or filename.endswith(".txt"):
            import csv, io as _io
            text     = content.decode("utf-8-sig")
            sample   = text[:4096]
            delim    = "\t" if sample.count("\t") > sample.count(",") else ","
            reader   = csv.DictReader(_io.StringIO(text), delimiter=delim)
            raw_rows = [dict(row) for row in reader]

        else:
            # Default: CSV
            import csv, io as _io
            text     = content.decode("utf-8-sig")
            reader   = csv.DictReader(_io.StringIO(text))
            raw_rows = [dict(row) for row in reader]

    except Exception as e:
        return {"error": f"Could not parse file: {str(e)}"}

    # ── column normalisation ───────────────────────────────────────────────
    CMAP = {
        "id":"experiment_id","exp_id":"experiment_id",
        "sample_id":"experiment_id","specimen_id":"experiment_id",
        "name":"experiment_name","material":"material_type",
        "mat":"material_type","alloy":"material_type",
        "batch":"material_batch","layer_thickness":"layer_height",
        "scan_speed":"print_speed","laser_speed":"print_speed",
        "nozzle_temp":"nozzle_temperature",
        "extrusion_temp":"nozzle_temperature",
        "platform_temp":"bed_temperature",
        "infill":"infill_percentage","tensile":"tensile_strength_mpa",
        "uts":"tensile_strength_mpa","roughness":"surface_roughness_um",
        "ra":"surface_roughness_um","porosity":"porosity_percent",
        "density":"density_g_per_cm3","elongation":"elongation_percent",
        "hardness":"hardness_hb","defects":"defect_count",
    }

    def norm(row):
        out = {}
        for k, v in row.items():
            clean = str(k).strip().lower().replace(" ","_").replace("-","_")
            out[CMAP.get(clean, clean)] = v
        return out

    def pf(v):
        try:
            return float(v) if str(v).strip() not in (
                "","None","N/A","nan","NaN","-","null") else None
        except: return None

    def pi(v):
        try:
            return int(float(v)) if str(v).strip() not in (
                "","None","N/A") else None
        except: return None

    # ── insert rows ────────────────────────────────────────────────────────
    created = skipped = failed = 0
    errors  = []

    for idx, raw in enumerate(raw_rows, 1):
        row    = norm(raw)
        exp_id = str(row.get("experiment_id") or "").strip() or f"UP-{idx:05d}"

        existing = db.query(Experiment).filter(
            Experiment.experiment_id == exp_id
        ).first()
        if existing:
            skipped += 1
            continue

        try:
            db.add(Experiment(
                experiment_id   = exp_id,
                experiment_name = str(row.get("experiment_name") or f"Upload {idx}"),
                material_type   = row.get("material_type")  or None,
                material_batch  = row.get("material_batch") or None,
                build_platform  = row.get("build_platform") or None,
                build_date      = row.get("build_date")      or None,
                operator        = row.get("operator")        or None,
                status          = row.get("status")          or "completed",
                notes           = row.get("notes")           or None,
            ))
            db.add(ProcessParameter(
                experiment_id     = exp_id,
                layer_height      = pf(row.get("layer_height")),
                print_speed       = pf(row.get("print_speed")),
                nozzle_temperature= pf(row.get("nozzle_temperature")),
                bed_temperature   = pf(row.get("bed_temperature")),
                infill_percentage = pf(row.get("infill_percentage")),
                infill_pattern    = row.get("infill_pattern") or None,
                shell_count       = pi(row.get("shell_count")),
                print_time_hours  = pf(row.get("print_time_hours")),
            ))
            db.add(GeometryData(
                experiment_id    = exp_id,
                volume_mm3       = pf(row.get("volume_mm3")),
                surface_area_mm2 = pf(row.get("surface_area_mm2")),
                bounding_box_x   = pf(row.get("bounding_box_x")),
                bounding_box_y   = pf(row.get("bounding_box_y")),
                bounding_box_z   = pf(row.get("bounding_box_z")),
            ))
            db.add(QualityMetric(
                experiment_id       = exp_id,
                tensile_strength_mpa= pf(row.get("tensile_strength_mpa")),
                yield_strength_mpa  = pf(row.get("yield_strength_mpa")),
                elongation_percent  = pf(row.get("elongation_percent")),
                surface_roughness_um= pf(row.get("surface_roughness_um")),
                porosity_percent    = pf(row.get("porosity_percent")),
                density_g_per_cm3   = pf(row.get("density_g_per_cm3")),
                hardness_hb         = pf(row.get("hardness_hb")),
                defect_count        = pi(row.get("defect_count")),
            ))
            db.commit()
            compute_ml_features(db, exp_id)
            created += 1
        except Exception as e:
            db.rollback()
            failed += 1
            errors.append({"row": idx, "id": exp_id, "error": str(e)[:120]})

    return {
        "filename":   file.filename,
        "format":     os.path.splitext(file.filename)[1].upper().lstrip("."), # type: ignore
        "total_rows": len(raw_rows),
        "created":    created,
        "skipped":    skipped,
        "failed":     failed,
        "errors":     errors[:10]
    }




# ── Step 7: startup banner ────────────────────────────────────────────────
if __name__ == "__main__":
    # Cloud hosts (Render, Railway, etc.) assign the port via the PORT env var.
    # Locally, if PORT is not set, this falls back to 8000 exactly as before.
    port = int(os.getenv("PORT", 8000))
    host = os.getenv("HOST", "0.0.0.0")

    # Show the public base URL if the host provides one, else localhost.
    base_url = os.getenv("RENDER_EXTERNAL_URL", f"http://localhost:{port}")

    print("=" * 60)
    print("  AM Data Pipeline — Unified Server")
    print("=" * 60)
    print(f"  Basic dashboard   : {base_url}")
    print(f"  Advanced dashboard: {base_url}/dashboard/advanced")
    print(f"  API docs (Swagger): {base_url}/api/v1/docs")
    print(f"  Health check      : {base_url}/api/v1/health")
    print(f"  Analytics JSON    : {base_url}/api/analytics/summary")
    print("=" * 60 + "\n")
    uvicorn.run(app, host=host, port=port, reload=False)