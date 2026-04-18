"""
load_local_dataset.py
Loads ANY locally saved dataset into the AM pipeline.

Supported formats:
  .csv     — comma separated values
  .tsv     — tab separated values
  .txt     — tab or comma separated text
  .xlsx    — Excel workbook (first sheet used)
  .xls     — Legacy Excel workbook
  .json    — JSON array of objects
  .parquet — Apache Parquet columnar format

Usage:
  python load_local_dataset.py                          # auto-detect file
  python load_local_dataset.py mydata.xlsx              # Excel file
  python load_local_dataset.py nist_data.csv            # CSV file
  python load_local_dataset.py "C:/Downloads/data.json" # full path
  python load_local_dataset.py results.parquet          # Parquet file
"""

import sys
import os
import time
import requests

API_BASE = "http://localhost:8000/api/v1"

# ── column name mappings ───────────────────────────────────────────────────
COLUMN_MAP = {
    "id":                    "experiment_id",
    "exp_id":                "experiment_id",
    "sample_id":             "experiment_id",
    "specimen_id":           "experiment_id",
    "name":                  "experiment_name",
    "exp_name":              "experiment_name",
    "sample_name":           "experiment_name",
    "material":              "material_type",
    "mat":                   "material_type",
    "alloy":                 "material_type",
    "polymer":               "material_type",
    "batch":                 "material_batch",
    "layer_thickness":       "layer_height",
    "layer_thickness_mm":    "layer_height",
    "hatch_spacing":         "print_speed",
    "scan_speed":            "print_speed",
    "laser_speed":           "print_speed",
    "laser_scan_speed":      "print_speed",
    "nozzle_temp":           "nozzle_temperature",
    "extrusion_temp":        "nozzle_temperature",
    "melt_temperature":      "nozzle_temperature",
    "platform_temp":         "bed_temperature",
    "build_plate_temp":      "bed_temperature",
    "preheat_temp":          "bed_temperature",
    "infill":                "infill_percentage",
    "fill_density":          "infill_percentage",
    "raster_angle":          "infill_pattern",
    "tensile":               "tensile_strength_mpa",
    "uts":                   "tensile_strength_mpa",
    "ultimate_tensile":      "tensile_strength_mpa",
    "tensile_mpa":           "tensile_strength_mpa",
    "uts_mpa":               "tensile_strength_mpa",
    "yield":                 "yield_strength_mpa",
    "ys":                    "yield_strength_mpa",
    "yield_mpa":             "yield_strength_mpa",
    "roughness":             "surface_roughness_um",
    "ra":                    "surface_roughness_um",
    "surface_ra":            "surface_roughness_um",
    "ra_um":                 "surface_roughness_um",
    "porosity":              "porosity_percent",
    "pore_fraction":         "porosity_percent",
    "relative_density":      "density_g_per_cm3",
    "density":               "density_g_per_cm3",
    "elongation":            "elongation_percent",
    "elongation_at_break":   "elongation_percent",
    "hardness":              "hardness_hb",
    "hardness_hv":           "hardness_hb",
    "defects":               "defect_count",
    "number_of_defects":     "defect_count",
}

# ── format readers ─────────────────────────────────────────────────────────

def read_csv(filepath: str) -> list:
    """Read a comma-separated CSV file."""
    import csv
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]

def read_tsv(filepath: str) -> list:
    """Read a tab-separated file (.tsv or .txt)."""
    import csv
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        # Auto-detect delimiter
        sample = f.read(4096)
        f.seek(0)
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(row) for row in reader]

def read_excel(filepath: str) -> list:
    """
    Read an Excel file (.xlsx or .xls).
    Uses the first sheet by default.
    If multiple sheets exist, lists them and asks which to use.
    """
    import openpyxl

    if filepath.endswith(".xls"):
        # Legacy format — convert via xlrd
        try:
            import xlrd
            wb   = xlrd.open_workbook(filepath)
            sheets = wb.sheet_names()
            print(f"   Excel sheets found: {sheets}")
            sheet_name = sheets[0]
            if len(sheets) > 1:
                choice = input(
                    f"   Multiple sheets found. Enter sheet name "
                    f"(or press Enter for '{sheets[0]}'): "
                ).strip()
                if choice and choice in sheets:
                    sheet_name = choice
            sheet  = wb.sheet_by_name(sheet_name)
            header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            rows   = []
            for r in range(1, sheet.nrows):
                row = {}
                for c in range(sheet.ncols):
                    val = sheet.cell_value(r, c)
                    row[header[c]] = str(val) if val != "" else ""
                rows.append(row)
            return rows
        except ImportError:
            print("   xlrd not installed. Run: pip install xlrd")
            raise

    # .xlsx via openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    print(f"   Excel sheets found: {sheet_names}")

    sheet_name = sheet_names[0]
    if len(sheet_names) > 1:
        choice = input(
            f"   Multiple sheets found. Enter sheet name "
            f"(or press Enter for '{sheet_names[0]}'): "
        ).strip()
        if choice and choice in sheet_names:
            sheet_name = choice

    ws     = wb[sheet_name]
    rows   = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c) if c is not None else f"col_{i}"
              for i, c in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue   # skip fully empty rows
        result.append({
            header[i]: (str(v) if v is not None else "")
            for i, v in enumerate(row)
        })
    return result

def read_json(filepath: str) -> list:
    """
    Read a JSON file.
    Supports:
      - Array of objects: [{...}, {...}]
      - Single object:    {...}  (wrapped in a list)
      - Nested key:       {"experiments": [{...}]}
    """
    import json
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return [
            {k: (str(v) if v is not None else "") for k, v in item.items()}
            for item in data
            if isinstance(item, dict)
        ]

    if isinstance(data, dict):
        # Look for the first list value in the dict
        for key, val in data.items():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                print(f"   JSON key used: '{key}'")
                return [
                    {k: (str(v) if v is not None else "") for k, v in item.items()}
                    for item in val
                ]
        # Single object — wrap as one-row list
        return [{k: (str(v) if v is not None else "") for k, v in data.items()}]

    raise ValueError("Unrecognised JSON structure. Expected array or object.")

def read_parquet(filepath: str) -> list:
    """Read an Apache Parquet file."""
    try:
        import pandas as pd
        df = pd.read_parquet(filepath)
        # Replace NaN with empty string for consistent handling
        df = df.where(df.notna(), other="")
        return df.astype(str).to_dict(orient="records")
    except ImportError:
        print("   pandas/pyarrow not installed. Run: pip install pandas pyarrow")
        raise

# ── format dispatcher ──────────────────────────────────────────────────────

FILE_READERS = {
    ".csv":     read_csv,
    ".tsv":     read_tsv,
    ".txt":     read_tsv,      # treat .txt as tsv (auto-detects delimiter)
    ".xlsx":    read_excel,
    ".xls":     read_excel,
    ".json":    read_json,
    ".parquet": read_parquet,
}

def load_file(filepath: str) -> list:
    """Detect file format and read rows."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in FILE_READERS:
        supported = ", ".join(FILE_READERS.keys())
        raise ValueError(
            f"Unsupported file type '{ext}'.\n"
            f"Supported formats: {supported}"
        )
    print(f"   Format detected : {ext.upper().lstrip('.')}")
    reader = FILE_READERS[ext]
    return reader(filepath)

# ── column normaliser ──────────────────────────────────────────────────────

def normalize_columns(row: dict) -> dict:
    """Remap column names to standard API field names."""
    normalized = {}
    for key, value in row.items():
        clean_key = (
            str(key).strip().lower()
            .replace(" ", "_").replace("-", "_").replace(".", "_")
        )
        mapped_key = COLUMN_MAP.get(clean_key, clean_key)
        normalized[mapped_key] = (
            value.strip() if isinstance(value, str) else value
        )
    return normalized

# ── value parsers ──────────────────────────────────────────────────────────

def parse_float(val):
    try:
        return float(val) if str(val).strip() not in (
            "", "None", "N/A", "NA", "nan", "NaN", "-", "null"
        ) else None
    except (ValueError, TypeError):
        return None

def parse_int(val):
    try:
        return int(float(val)) if str(val).strip() not in (
            "", "None", "N/A", "NA"
        ) else None
    except (ValueError, TypeError):
        return None

def parse_bool(val):
    return str(val).strip().lower() in ("true", "1", "yes", "y")

# ── payload builder ────────────────────────────────────────────────────────

def row_to_payload(row: dict, index: int) -> dict:
    """Convert a normalised row dict into the API request body."""
    exp_id = str(row.get("experiment_id") or "").strip()
    if not exp_id:
        mat    = str(row.get("material_type", "EXP"))[:3].upper()
        exp_id = f"{mat}-LOCAL-{index:04d}"

    exp_name = str(row.get("experiment_name") or "").strip()
    if not exp_name:
        material = row.get("material_type", "Unknown")
        layer    = row.get("layer_height", "")
        exp_name = (
            f"{material} experiment {index}"
            + (f" layer {layer}mm" if layer else "")
        )

    return {
        "experiment_id":   exp_id,
        "experiment_name": exp_name,
        "material_type":   row.get("material_type")  or None,
        "material_batch":  row.get("material_batch") or None,
        "build_platform":  row.get("build_platform") or None,
        "build_date":      row.get("build_date")      or None,
        "operator":        row.get("operator")        or None,
        "status":          row.get("status")          or "completed",
        "notes":           row.get("notes")           or None,
        "process_parameters": {
            "layer_height":       parse_float(row.get("layer_height")),
            "print_speed":        parse_float(row.get("print_speed")),
            "nozzle_temperature": parse_float(row.get("nozzle_temperature")),
            "bed_temperature":    parse_float(row.get("bed_temperature")),
            "infill_percentage":  parse_float(row.get("infill_percentage")),
            "infill_pattern":     row.get("infill_pattern") or None,
            "shell_count":        parse_int(row.get("shell_count")),
            "support_enabled":    parse_bool(row.get("support_enabled")),
            "print_time_hours":   parse_float(row.get("print_time_hours")),
        },
        "geometry_data": {
            "volume_mm3":        parse_float(row.get("volume_mm3")),
            "surface_area_mm2":  parse_float(row.get("surface_area_mm2")),
            "bounding_box_x":    parse_float(row.get("bounding_box_x")),
            "bounding_box_y":    parse_float(row.get("bounding_box_y")),
            "bounding_box_z":    parse_float(row.get("bounding_box_z")),
        },
        "quality_metrics": {
            "tensile_strength_mpa": parse_float(row.get("tensile_strength_mpa")),
            "yield_strength_mpa":   parse_float(row.get("yield_strength_mpa")),
            "elongation_percent":   parse_float(row.get("elongation_percent")),
            "surface_roughness_um": parse_float(row.get("surface_roughness_um")),
            "porosity_percent":     parse_float(row.get("porosity_percent")),
            "density_g_per_cm3":    parse_float(row.get("density_g_per_cm3")),
            "hardness_hb":          parse_float(row.get("hardness_hb")),
            "defect_count":         parse_int(row.get("defect_count")),
        }
    }

# ── column preview ─────────────────────────────────────────────────────────

def preview_columns(rows: list, filepath: str):
    """Print a summary of what columns were found."""
    if not rows:
        print("   No rows found in file.")
        return

    cols = list(rows[0].keys())
    print(f"\n   File       : {filepath}")
    print(f"   Total rows : {len(rows)}")
    print(f"   Columns    : {len(cols)}")
    print()
    print(f"   {'Original column':<35} {'Mapped to':<30} {'Status'}")
    print(f"   {'-'*35} {'-'*30} {'-'*10}")

    api_fields = set(COLUMN_MAP.values()) | {
        "experiment_id", "experiment_name", "material_type",
        "layer_height", "nozzle_temperature", "tensile_strength_mpa",
        "porosity_percent", "density_g_per_cm3", "surface_roughness_um"
    }

    for col in cols:
        clean  = col.strip().lower().replace(" ", "_").replace("-","_")
        mapped = COLUMN_MAP.get(clean, clean)
        known  = mapped in api_fields
        status = "mapped" if mapped != clean else ("known" if known else "extra")
        print(f"   {col:<35} {mapped:<30} {status}")
    print()

# ── main ───────────────────────────────────────────────────────────────────

def main():
    # Determine input file
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        # Auto-detect common filenames
        candidates = [
            "nist_inspired_am_dataset.csv",
            "am_bulk_dataset.csv",
            "am_dataset.csv",
            "nist_data.xlsx",
            "am_data.xlsx",
            "dataset.json",
            "dataset.parquet",
            "data.csv",
        ]
        filepath = next((c for c in candidates if os.path.exists(c)), None)
        if not filepath:
            print("No dataset file found in current folder.\n")
            print("Usage:")
            print("  python load_local_dataset.py yourfile.csv")
            print("  python load_local_dataset.py yourfile.xlsx")
            print("  python load_local_dataset.py yourfile.json")
            print("  python load_local_dataset.py yourfile.parquet")
            print("  python load_local_dataset.py yourfile.tsv")
            print("\nSupported: .csv  .tsv  .txt  .xlsx  .xls  .json  .parquet")
            return

    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return

    print("=" * 60)
    print("  MULTI-FORMAT DATASET LOADER")
    print("=" * 60)

    # Load file
    print(f"\nReading file: {filepath}")
    try:
        rows = load_file(filepath)
    except Exception as e:
        print(f"Failed to read file: {e}")
        return

    if not rows:
        print("File is empty or could not be parsed.")
        return

    # Normalise columns
    rows = [normalize_columns(r) for r in rows]

    # Preview
    preview_columns(rows, filepath)

    # Confirm
    confirm = input("Proceed with upload? (yes/no): ").strip().lower()
    if confirm not in ("yes", "y"):
        print("Cancelled.")
        return

    # Verify server
    try:
        r = requests.get(f"{API_BASE}/health", timeout=5)
        print(f"\nServer: {r.json().get('status')}\n")
    except Exception as e:
        print(f"Server not reachable: {e}")
        print("Start the server first: python server.py")
        return

    # Upload
    total   = len(rows)
    created = skipped = failed = 0
    errors  = []
    start   = time.time()

    print(f"Uploading {total} rows...")
    print("-" * 55)

    for idx, row in enumerate(rows, 1):
        try:
            payload  = row_to_payload(row, idx)
            response = requests.post(
                f"{API_BASE}/experiments",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=15
            )
            if response.status_code == 201:
                created += 1
                if idx % 10 == 0 or idx == total:
                    print(f"  Progress: {idx}/{total} — "
                          f"{created} created, {skipped} skipped, "
                          f"{failed} failed")
            elif response.status_code == 409:
                skipped += 1
            else:
                failed += 1
                detail = response.json().get("detail", response.text[:80])
                errors.append(f"Row {idx}: {detail}")
                print(f"  FAILED row {idx}: {detail}")
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx}: {str(e)}")

    elapsed = round(time.time() - start, 2)

    print(f"\n{'='*55}")
    print("UPLOAD COMPLETE")
    print(f"{'='*55}")
    print(f"  File     : {os.path.basename(filepath)}")
    print(f"  Format   : {os.path.splitext(filepath)[1].upper().lstrip('.')}")
    print(f"  Created  : {created}")
    print(f"  Skipped  : {skipped}  (already in DB)")
    print(f"  Failed   : {failed}")
    print(f"  Time     : {elapsed}s")
    if errors:
        print(f"\n  First errors:")
        for e in errors[:5]:
            print(f"    {e}")
    print(f"\n  Dashboard : http://localhost:8000")
    print(f"  Analytics : http://localhost:8000/api/analytics/summary")
    print("=" * 55)

if __name__ == "__main__":
    main()